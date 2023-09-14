from openpyxl import load_workbook, Workbook
from docxtpl import DocxTemplate
from os.path import join, abspath
import csv
import os

GRADE = 6
NUM_COLUMN = 3
FILE_CODES = 'sch779841_6.csv'

def create_file(file_input, context, file_output):
    doc = DocxTemplate(file_input)
    doc.render(context)
    doc.save(file_output)

# Открытие файла Excel и загрузка рабочей книги в переменную
data_path = join('.', 'students.xlsx')
data_path = abspath(data_path)
wb = load_workbook(filename=data_path, data_only=True, read_only=True)
sheet = wb.active

# Получение списка ФИО студентов из 2 столбца таблицы Excel
students = []
for i in range(2, sheet.max_row + 1):
    if sheet.cell(row = i, column = 2).value == GRADE:
        dict_student = {
            'name': sheet.cell(row = i, column = 1).value,
            'class': sheet.cell(row = i, column = 2).value,
            'letter': sheet.cell(row = i, column = 3).value,
        }
        students.append(dict_student)

# Открытие файла CSV и загрузка данный в список
list_student = []
with open(FILE_CODES, encoding='cp1251') as csvfile:
    codes = list(csv.reader(csvfile, delimiter=';'))

FOLDER_REPORT = f"{GRADE} классы"
if not os.path.exists(FOLDER_REPORT):
    os.mkdir(FOLDER_REPORT) 

# Формирование списка словарей на основе полученных данных
name_columns = codes[0]
name_subjects = name_columns[NUM_COLUMN:]
for i, row in enumerate(codes):
    temp_dict = {name:value for name, value in zip(name_columns, row)}
    if len(students) == i:
        break
    temp_dict['ФИО'] = students[i]['name']
    subjects_olympiad = []
    for j, code in enumerate(row[NUM_COLUMN:]):
        subjects_olympiad.append({
            'subject': name_subjects[j],
            'code': code,
            'grades': '-',
            'dates': '-',
            })
    context = {
        'student_name': students[i]['name'],
        'student_group': f"{students[i]['class']}{students[i]['letter']}",
        'dataOlympiads': subjects_olympiad,
    }
    create_file("template.docx", context, os.path.join(FOLDER_REPORT, f"{students[i]['class']}{students[i]['letter']} {students[i]['name']}.docx"))
    list_student.append(temp_dict)

# Экспорт таблицы со студентами в Excel
wb_export = Workbook()
ws_export = wb_export.active
ws_export.append(name_columns)
for student in list_student:
    values = (student[k] for k in name_columns)
    ws_export.append(values)
wb_export.save(filename='list.xlsx')