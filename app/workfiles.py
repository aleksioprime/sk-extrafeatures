from app import app
from docx import Document
from io import BytesIO
import json
from docxtpl import DocxTemplate
from openpyxl import load_workbook, Workbook
from flask import send_file, request, jsonify, session
from fileinput import filename
import csv
import os
from datetime import datetime, date
import zipfile

def create_file(file_input, context):
    document = DocxTemplate(file_input)
    document.render(context)
    return document
    

@app.route('/download/marks', methods=["POST"])
def download_marks():
    data_student = json.loads(request.data)['data']
    subjects = {}
    for group in data_student['groups']:
       for key, value in group['marks'].items():
           subjects[str(key)] = value['subjectName']
           group['marks'][key] = value['mark']
    document = create_file("doc_templates/template_marks.docx", { 'groups': data_student['groups'], 'subjects': subjects, 'student_name': data_student['student']['shortName']})
    buffer = BytesIO()
    document.save(buffer)
    buffer.seek(0)
    return send_file(buffer,  download_name='report.docx')


@app.route('/download/codes', methods=["POST"])
def download_codes():
    data_codes = json.loads(request.data)['codes']
    data_subjects = json.loads(request.data)['subjects']
    data_grade = json.loads(request.data)['grade']
    # Экспорт таблицы со студентами в Excel
    wb_export = Workbook()
    ws_export = wb_export.active
    ws_export.column_dimensions['A'].width = 30
    ws_export.column_dimensions['B'].width = 5
    name_columns = ['Студент', 'Класс'] + [data['name'] for data in data_subjects]
    ws_export.append(name_columns)
    print(name_columns)
    date_subjects = ['', ''] + [data['date'] for data in data_subjects]
    ws_export.append(date_subjects)
    for student in data_codes:
        values = (student[k] for k in name_columns)
        ws_export.append(values)
    buffer = BytesIO()
    wb_export.save(buffer)
    buffer.seek(0)
    return send_file(buffer,  download_name=f'{data_grade} grade codes ({date.today()}).xlsx')

@app.route('/download/student/codes', methods=["POST"])
def download_student_codes():
    data_codes = json.loads(request.data)['codes']
    data_subjects = json.loads(request.data)['subjects']
    data_grade = json.loads(request.data)['grade']
    data_school = json.loads(request.data)['school']
    print(data_subjects)
    
    temp_path = os.path.join('doc_templates', 'student_codes', f"{data_grade}_class")
    if not os.path.exists(temp_path):
        os.mkdir(temp_path)
    for data_student in data_codes:
        data_code_subjects = [{ 'name': subject['name'], 'date': subject['date'], 'code': data_student[subject['name']] } for subject in data_subjects]
        context = {
            'student_name': data_student['Студент'],
            'student_group': data_student['Класс'],
            'data_subjects': data_code_subjects,
            'school_login': data_school
        }
        document = create_file("doc_templates/template_studentcode.docx", context)
        document.save(os.path.join(temp_path, f"{data_student['Класс']}-{data_student['Студент']}.docx"))

    memory_file = BytesIO()
    with zipfile.ZipFile(memory_file, 'w') as zf:
        for file in os.listdir(temp_path):
            zf.write(os.path.join(temp_path, file), os.path.basename(file), compress_type=zipfile.ZIP_DEFLATED)
    memory_file.seek(0)

    for f in os.listdir(temp_path):
        os.remove(os.path.join(temp_path, f))
    os.rmdir(temp_path)

    return send_file(memory_file,  download_name=f'{data_grade} grade reports ({date.today()}).zip')

@app.route('/upload/codes', methods=["POST"])
def upload_codes():
    file_codes = request.files['file_codes']
    # print(json.loads(request.form['students'])[0])
    students = json.loads(request.form['students'])
    if file_codes:
        if not os.path.exists('upload_files'):
            os.mkdir('upload_files')
        file_name = os.path.join('upload_files', f"{session['email']}-{datetime.now()}-{file_codes.filename}")
        file_codes.save(file_name)
        # file_codes.seek(0)
        # text_file = file_codes.read().decode('cp1251')
        # codes = [[ cell for cell in line.split(';') ] for line in text_file.split('\r\n')]
        # name_columns = codes[2]
        # name_subjects = name_columns[3:]
        # print(name_subjects)

        codes = []
        count_students = 0
        school_codes = []
        school = 'Not found'
        with open(file_name, encoding="cp1251") as csvfile:
            reader = csv.reader(csvfile, delimiter=';',)  
            for index, row in enumerate(reader):
                if index == 0:
                    name_subjects = row[4:]
                elif index == 1:
                    dates = [ { 'name': name, 'date': date } for name, date in zip(name_subjects, row[4:])]
                else:
                    dict_codes = {name:value for name, value in zip(name_subjects, row[4:])}
                    dict_codes['Студент'] = students[count_students]['name']
                    dict_codes['Класс'] = f"{students[count_students]['class']}{students[count_students]['letter']}"
                    school_codes.append(row[0])
                    codes.append(dict_codes)
                    count_students += 1
                    if count_students >= len(students):
                        break
            if len(set(school_codes)) == 1:
                school = school_codes[0]
            
        os.remove(file_name)

        return jsonify({
            'status': 'success',
            'codes': codes,
            'subjects': dates,
            'school': school,
        })
    else:
        return jsonify("Failed")
    
@app.route('/upload/students', methods=["POST"])
def upload_students():
    file_students = request.files['file_students']
    if file_students:
        if not os.path.exists('upload_files'):
            os.mkdir('upload_files')
        file_name = os.path.join('upload_files', f"{session['email']}-{datetime.now()}-{file_students.filename}")
        file_students.save(file_name)
        wb = load_workbook(filename=file_name, data_only=True, read_only=True)
        sheet = wb.active
        # Получение списка ФИО студентов из 2 столбца таблицы Excel
        students = []
        for i in range(2, sheet.max_row + 1):
            students.append({
                'name': sheet.cell(row = i, column = 1).value,
                'class': sheet.cell(row = i, column = 2).value,
                'letter': sheet.cell(row = i, column = 3).value,
            })

        os.remove(file_name)

        return jsonify({
            'status': 'success',
            'students': students,
        })
    else:
        return jsonify("Failed")